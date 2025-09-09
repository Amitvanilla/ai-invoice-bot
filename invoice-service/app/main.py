"""
Invoice Processing Service - FastAPI Application
AI-powered invoice parsing, classification, and semantic search
OPTIMIZED AI FLOW:
1. Landing AI - Extracts raw data (markdown + chunks) from invoice PDFs
2. Claude 3-7-sonnet - Converts Landing AI results to structured invoice data format
3. Azure OpenAI - Validates and corrects extraction against raw data
"""

import structlog
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.orm import Session
from typing import List, Optional
import uvicorn
import os
import json
import uuid
from datetime import datetime
from dotenv import load_dotenv
import anthropic
from openai import AzureOpenAI
from tenacity import retry, stop_after_attempt, wait_exponential
import numpy as np
import gc
import logging
import psycopg2
from psycopg2.extras import RealDictCursor

# Handle Excel import with fallback
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError as e:
    print(f"Warning: pandas/openpyxl not available: {e}")
    EXCEL_AVAILABLE = False

load_dotenv()

# Set environment variables
os.environ['VISION_AGENT_API_KEY'] = os.getenv('VISION_AGENT_API_KEY')
os.environ['ANTHROPIC_API_KEY'] = os.getenv('ANTHROPIC_API_KEY')

# Azure OpenAI configuration
AZURE_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_DEPLOYMENT = os.getenv("AZURE_DEPLOYMENT_NAME")
AZURE_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")

# Custom JSON encoder to handle NumPy types and Chunk objects
class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif hasattr(obj, '__dict__'):
            # Handle objects with __dict__ (like Chunk objects)
            try:
                return {k: v for k, v in obj.__dict__.items()
                        if not k.startswith('_')}
            except Exception:
                return str(obj)
        elif hasattr(obj, 'to_dict'):
            # Handle objects with to_dict method
            try:
                return obj.to_dict()
            except Exception:
                return str(obj)
        else:
            # Fallback to string representation
            try:
                return str(obj)
            except Exception:
                return f"<{type(obj).__name__} object>"

from .core.config import settings
from .core.security import verify_token
from .db.session import get_db
from .db.models import Invoice
from .schemas.invoice import InvoiceResponse, InvoiceQuery, InvoiceSearchResponse
from .services.invoice_processor import InvoiceProcessor
from .services.embeddings import EmbeddingService
from .services.search import SearchService
from .api.routes import invoices, auth

def update_invoice_database_directly(invoice_id: str, extracted_data: dict, filename: str):
    """
    Directly update the database with extracted data using psycopg2
    """
    try:
        print(f"\n🔍 DEBUG: Starting database update for invoice: {invoice_id}")
        print(f"🔍 DEBUG: Input extracted_data keys: {list(extracted_data.keys()) if extracted_data else 'None'}")
        print(f"🔍 DEBUG: Input extracted_data type: {type(extracted_data)}")
        
        # Extract the key data from the processed results
        vendor_name = "Unknown Vendor"
        total_amount = "0"
        invoice_number = "Unknown"
        
        # Extract vendor name from CSV data
        print(f"🔍 DEBUG: Checking vendor_info...")
        if extracted_data.get("vendor_info") and len(extracted_data["vendor_info"]) > 0:
            vendor_data = extracted_data["vendor_info"][0].get("data", "")
            print(f"🔍 DEBUG: Vendor data: {vendor_data[:200]}...")
            if "Vendor Name," in vendor_data:
                vendor_match = vendor_data.split("Vendor Name,")[1].split("\n")[0].strip()
                print(f"🔍 DEBUG: Vendor match: '{vendor_match}'")
                if vendor_match and vendor_match != "UNKNOWN":
                    vendor_name = vendor_match
                    print(f"✅ DEBUG: Set vendor_name to: {vendor_name}")
        else:
            print(f"🔍 DEBUG: No vendor_info found or empty")
        
        # Extract total amount from CSV data
        print(f"🔍 DEBUG: Checking payment_info...")
        if extracted_data.get("payment_info") and len(extracted_data["payment_info"]) > 0:
            payment_data = extracted_data["payment_info"][0].get("data", "")
            print(f"🔍 DEBUG: Payment data: {payment_data[:200]}...")
            if "Total Amount Due," in payment_data:
                amount_match = payment_data.split("Total Amount Due,")[1].split("\n")[0].strip()
                print(f"🔍 DEBUG: Amount match: '{amount_match}'")
                if amount_match and amount_match != "UNKNOWN":
                    total_amount = amount_match
                    print(f"✅ DEBUG: Set total_amount to: {total_amount}")
        else:
            print(f"🔍 DEBUG: No payment_info found or empty")
        
        # Extract invoice number from CSV data
        print(f"🔍 DEBUG: Checking invoice_details...")
        if extracted_data.get("invoice_details") and len(extracted_data["invoice_details"]) > 0:
            invoice_data = extracted_data["invoice_details"][0].get("data", "")
            print(f"🔍 DEBUG: Invoice data: {invoice_data[:200]}...")
            if "Invoice Number," in invoice_data:
                invoice_match = invoice_data.split("Invoice Number,")[1].split("\n")[0].strip()
                print(f"🔍 DEBUG: Invoice match: '{invoice_match}'")
                if invoice_match and invoice_match != "UNKNOWN":
                    invoice_number = invoice_match
                    print(f"✅ DEBUG: Set invoice_number to: {invoice_number}")
        else:
            print(f"🔍 DEBUG: No invoice_details found or empty")
        
        print(f"🔄 Updating database directly for invoice: {invoice_id}")
        print(f"📊 Final extracted data - Vendor: {vendor_name}, Amount: {total_amount}, Invoice #: {invoice_number}")
        
        # Connect to the database using the full URL
        from .core.config import settings
        conn = psycopg2.connect(settings.DATABASE_URL)
        
        with conn.cursor() as cursor:
            # First, get the current extracted_data to preserve other fields
            cursor.execute("SELECT \"extractedData\", \"embeddings\" FROM \"Invoice\" WHERE id = %s", (invoice_id,))
            result = cursor.fetchone()
            
            if result:
                current_data = result[0] if result[0] else {}
                current_embeddings = result[1] if result[1] else None
                
                # Extract embeddings from current_data if present and move to embeddings field
                embeddings_json = current_embeddings
                if current_data.get("embeddings") and isinstance(current_data["embeddings"], list):
                    embeddings_json = json.dumps(current_data["embeddings"])
                    # Remove embeddings from extracted_data
                    current_data = {k: v for k, v in current_data.items() if k != "embeddings"}
                
                # Update the extracted_data with new values
                updated_data = {
                    **current_data,
                    "vendor_name": vendor_name,
                    "total_amount": total_amount,
                    "invoice_number": invoice_number,
                    "file_processed": True,
                    "excel_generated": True,
                    "status": "Successfully processed by AI"
                }
                
                # Update the database with both extractedData and embeddings
                if embeddings_json:
                    cursor.execute(
                        "UPDATE \"Invoice\" SET \"extractedData\" = %s, \"embeddings\" = %s, status = %s WHERE id = %s",
                        (json.dumps(updated_data), embeddings_json, "processed", invoice_id)
                    )
                else:
                    cursor.execute(
                        "UPDATE \"Invoice\" SET \"extractedData\" = %s, status = %s WHERE id = %s",
                        (json.dumps(updated_data), "processed", invoice_id)
                    )
                
                conn.commit()
                print(f"✅ Successfully updated database directly for invoice {invoice_id}")
            else:
                print(f"❌ Invoice {invoice_id} not found in database")
                
    except Exception as e:
        print(f"❌ Error updating database directly: {e}")
        # Don't raise the exception to avoid breaking the main processing flow
    finally:
        if 'conn' in locals():
            conn.close()

# Global variable to store last processed file
last_processed_file = None

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure structured logging
structlog.configure(
    processors=[
        structlog.stdlib.filter_by_level,
        structlog.stdlib.add_logger_name,
        structlog.stdlib.add_log_level,
        structlog.stdlib.PositionalArgumentsFormatter(),
        structlog.processors.TimeStamper(fmt="iso"),
        structlog.processors.StackInfoRenderer(),
        structlog.processors.format_exc_info,
        structlog.processors.UnicodeDecoder(),
        structlog.processors.JSONRenderer()
    ],
    context_class=dict,
    logger_factory=structlog.stdlib.LoggerFactory(),
    wrapper_class=structlog.stdlib.BoundLogger,
    cache_logger_on_first_use=True,
)

# Initialize FastAPI app
app = FastAPI(
    title="Invoice Processing Service",
    description="AI-powered invoice parsing, classification, and semantic search",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# Configure CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Security
security = HTTPBearer()

# Initialize services
invoice_processor = InvoiceProcessor()
embedding_service = EmbeddingService()
search_service = SearchService()


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=4, max=10),
    reraise=True
)
def extract_invoice_data_with_landing_ai(pdf_path):
    """
    Use Landing AI Agentic Document Extraction to parse the entire invoice
    and extract comprehensive invoice information.
    """
    try:
        print(f"Extracting invoice data from PDF using Landing AI: {pdf_path}")

        # Import Landing AI library
        from agentic_doc.parse import parse

        # Create logs directory if it doesn't exist
        logs_dir = os.path.join(os.getcwd(), "logs")
        if not os.path.exists(logs_dir):
            os.makedirs(logs_dir)

        # Generate a unique file name for this extraction
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_name = os.path.basename(pdf_path)
        log_file_base = f"{timestamp}_{pdf_name}"

        # Check if PDF is encrypted before attempting to parse
        try:
            from pypdf import PdfReader
            reader = PdfReader(pdf_path)
            if reader.is_encrypted:
                raise Exception("PDF file is password-protected or encrypted. Cannot process encrypted PDFs.")
        except ImportError:
            print("Warning: pypdf not available for encryption check")
        except Exception as encrypt_error:
            if "encrypted" in str(encrypt_error).lower() or "password" in str(encrypt_error).lower():
                raise Exception("PDF file is password-protected or encrypted. Cannot process encrypted PDFs.")
            # Re-raise other PDF errors
            raise encrypt_error

        # Parse the PDF using Landing AI
        print("Calling Landing AI parse function...")
        results = parse(pdf_path)

        # Landing AI returns a list of results, handle properly
        if not results or len(results) == 0:
            print("No results returned from Landing AI")
            return {
                "vendor_info": [],
                "invoice_details": [],
                "line_items": [],
                "taxes_fees": [],
                "payment_info": [],
                "compliance_flags": []
            }

        # Take the first result (or combine multiple if needed)
        result = results[0] if len(results) == 1 else results

        # Get both markdown and structured data
        if hasattr(result, 'markdown') and hasattr(result, 'chunks'):
            # Single result object
            markdown_data = result.markdown
            chunks_data = result.chunks
        elif isinstance(results, list) and len(results) > 0:
            # Multiple results - combine them
            markdown_data = ""
            chunks_data = []
            for i, res in enumerate(results):
                if hasattr(res, 'markdown'):
                    markdown_data += f"\n--- Page {i+1} ---\n" + res.markdown
                if hasattr(res, 'chunks'):
                    chunks_data.extend(res.chunks)
        else:
            print("Unexpected result format from Landing AI")
            return {
                "vendor_info": [],
                "invoice_details": [],
                "line_items": [],
                "taxes_fees": [],
                "payment_info": [],
                "compliance_flags": []
            }

        # Save raw Landing AI results for debugging
        raw_result_file = os.path.join(logs_dir,
                                      f"{log_file_base}_landing_ai_raw.json")

        # Convert chunks to serializable format
        serializable_chunks = []
        if chunks_data:
            for chunk in chunks_data:
                try:
                    # Try to convert chunk to dict if it has __dict__
                    if hasattr(chunk, '__dict__'):
                        chunk_dict = {k: v for k, v in chunk.__dict__.items()
                                     if not k.startswith('_')}
                        serializable_chunks.append(chunk_dict)
                    else:
                        serializable_chunks.append(str(chunk))
                except Exception:
                    # Fallback to string representation
                    serializable_chunks.append(
                        f"<Chunk object: {str(chunk)[:200]}...>")

        with open(raw_result_file, 'w', encoding='utf-8') as f:
            json.dump({
                "results_count": len(results),
                "results_type": str(type(results)),
                "markdown": markdown_data,
                "chunks": serializable_chunks,
                "raw_results": [str(res) for res in results]
            }, f, indent=2, cls=NumpyEncoder)
        print(f"Saved raw Landing AI results to: {raw_result_file}")

        # Also save markdown to a separate file for easy reading
        markdown_file = os.path.join(logs_dir,
                                    f"{log_file_base}_landing_ai_markdown.md")
        with open(markdown_file, 'w', encoding='utf-8') as f:
            f.write(markdown_data)
        print(f"Saved Landing AI markdown to: {markdown_file}")

        # Convert Landing AI results to structured invoice data format
        result_data = convert_landing_ai_to_invoice_format(markdown_data,
                                                         chunks_data)

        # Store raw data in result for evaluator
        result_data['_raw_markdown'] = markdown_data
        result_data['_raw_chunks'] = chunks_data

        # Save processed data
        processed_data_file = os.path.join(logs_dir,
                                          f"{log_file_base}_processed_data.json")
        with open(processed_data_file, 'w', encoding='utf-8') as f:
            json.dump(result_data, f, indent=2, cls=NumpyEncoder, ensure_ascii=False)
        print(f"Saved processed data to: {processed_data_file}")

        return result_data

    except Exception as e:
        print(f"Error in Landing AI extraction: {e}")
        print(f"Error type: {type(e)}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")

        # Log the error with more detail
        try:
            logs_dir = os.path.join(os.getcwd(), "logs")
            if not os.path.exists(logs_dir):
                os.makedirs(logs_dir)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            pdf_name = os.path.basename(pdf_path) if pdf_path else "unknown"
            error_file = os.path.join(logs_dir,
                                     f"{timestamp}_{pdf_name}_error.txt")

            with open(error_file, 'w') as f:
                f.write(f"Error processing PDF with Landing AI: {str(e)}\n")
                f.write(f"Error type: {type(e)}\n")
                f.write(f"Traceback: {traceback.format_exc()}\n")
                f.write(f"PDF path: {pdf_path}\n")
        except Exception as log_error:
            print(f"Failed to log error: {log_error}")

        # Return empty but valid structure
        return {
            "vendor_info": [],
            "invoice_details": [],
            "line_items": [],
            "taxes_fees": [],
            "payment_info": [],
            "compliance_flags": []
        }


def convert_landing_ai_to_invoice_format(markdown_data, chunks_data):
    """
    Convert Landing AI results to structured invoice data format.
    This function analyzes the markdown and chunks data to extract invoice
    information.
    """
    try:
        print("Converting Landing AI results to invoice data format...")

        # Create logs directory for debugging
        logs_dir = os.path.join(os.getcwd(), "logs")
        if not os.path.exists(logs_dir):
            os.makedirs(logs_dir)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file_base = f"{timestamp}_invoice_conversion"

        # Log the input data for debugging
        input_debug_file = os.path.join(
            logs_dir, f"{log_file_base}_input_debug.json")
        with open(input_debug_file, 'w', encoding='utf-8') as f:
            json.dump({
                "markdown_length": len(markdown_data) if markdown_data else 0,
                "chunks_count": len(chunks_data) if chunks_data else 0,
                "markdown_preview": (markdown_data[:1000] if markdown_data
                                   else "No markdown data"),
                "chunks_preview": (chunks_data[:3] if chunks_data
                                 else "No chunks data")
            }, f, indent=2, cls=NumpyEncoder)
        print(f"Saved conversion input debug to: {input_debug_file}")

        # Use Claude to intelligently extract invoice data from Landing AI results
        client = anthropic.Anthropic()

        # Create comprehensive prompt to analyze invoice document data
        prompt = f"""Analyze the following invoice document data and extract
        comprehensive invoice information. This may include vendor invoices,
        purchase orders, receipts, or billing statements.

INVOICE MARKDOWN:
{markdown_data}

STRUCTURED CHUNKS:
{json.dumps(chunks_data, indent=2, cls=NumpyEncoder)}

Your task is to extract invoice information and identify any compliance or
processing flags:

1. VENDOR INFORMATION: Extract vendor/supplier details in CSV format:
   Field Name,Value
   REQUIRED FIELDS TO FIND:
   - Vendor Name (company/individual providing goods/services)
   - Vendor Address (complete address)
   - Vendor Contact (phone, email if available)
   - Tax ID (if available)
   - Payment Terms (net 30, due on receipt, etc.)

2. INVOICE DETAILS: Extract invoice header information in CSV format:
   Field Name,Value
   REQUIRED FIELDS TO FIND:
   - Invoice Number (unique identifier)
   - Invoice Date (when invoice was issued)
   - Due Date (payment due date)
   - Purchase Order Number (if referenced)
   - Currency (USD, EUR, etc.)

3. LINE ITEMS: Extract all line items in CSV format:
   Item Description,Quantity,Unit Price,Line Total
   REQUIRED FIELDS TO FIND:
   - Item descriptions
   - Quantities ordered/purchased
   - Unit prices
   - Extended amounts (quantity × unit price)

4. TAXES & FEES: Extract tax and fee information in CSV format:
   Tax/Fee Type,Rate,Amount
   REQUIRED FIELDS TO FIND:
   - Sales tax amounts and rates
   - Shipping/freight charges
   - Handling fees
   - Discounts applied

5. PAYMENT INFORMATION: Extract payment details in CSV format:
   Field Name,Value
   REQUIRED FIELDS TO FIND:
   - Total Amount Due
   - Amount Paid (if any)
   - Balance Due
   - Payment Method (if specified)

6. COMPLIANCE FLAGS: Extract compliance and processing flags in CSV format:
   Flag Type,Severity,Description
   REQUIRED FIELDS TO FIND:
   - Duplicate invoice detection
   - Amount discrepancies
   - Missing required fields
   - Tax compliance issues
   - Approval required flags

CRITICAL EXTRACTION RULES:
- Extract actual invoice data from invoices, receipts, and billing documents
- Handle negative amounts: If a value has a negative sign, preserve it.
  If a value is in parentheses, convert to negative
- Do NOT include instructions, explanations, or non-data text
- Extract actual invoice data (vendor names, invoice numbers, amounts, dates, etc.)
- Ensure all numeric values are clean (no commas in numbers, currency symbols removed)
- If any field cannot be determined, use "UNKNOWN"
- All required fields must be present
- Pay special attention to invoice numbers, dates, amounts, and vendor information

Respond in JSON format with these exact keys:
{{
  "vendor_info": [{{"data": "CSV data with headers"}}],
  "invoice_details": [{{"data": "CSV data with headers"}}],
  "line_items": [{{"data": "CSV data with headers"}}],
  "taxes_fees": [{{"data": "CSV data with headers"}}],
  "payment_info": [{{"data": "CSV data with headers"}}],
  "compliance_flags": [{{"data": "CSV data with headers"}}]
}}

Return ONLY valid JSON with NO explanatory text."""

        print("Calling Claude to convert Landing AI results...")
        response = client.messages.create(
            model="claude-3-7-sonnet-latest",
            max_tokens=8192,
            messages=[{
                "role": "user",
                "content": prompt
            }]
        )

        # Save Claude's response for debugging
        claude_response_file = os.path.join(logs_dir,
                                           f"{log_file_base}_claude_response.txt")
        with open(claude_response_file, 'w', encoding='utf-8') as f:
            f.write(response.content[0].text)
        print(f"Saved Claude response to: {claude_response_file}")

        # Parse Claude's response
        try:
            response_text = response.content[0].text.strip()

            # Strip markdown code blocks if present
            if response_text.startswith('```json'):
                response_text = response_text[7:]  # Remove ```json
            if response_text.startswith('```'):
                response_text = response_text[3:]   # Remove ```
            if response_text.endswith('```'):
                response_text = response_text[:-3]  # Remove closing ```

            # Clean up any remaining formatting
            response_text = response_text.strip()

            claude_result = json.loads(response_text)

            # Save parsed result for debugging
            claude_parsed_file = os.path.join(logs_dir,
                                             f"{log_file_base}_claude_parsed.json")
            with open(claude_parsed_file, 'w', encoding='utf-8') as f:
                json.dump(claude_result, f, indent=2, cls=NumpyEncoder, ensure_ascii=False)
            print(f"Saved Claude parsed result to: {claude_parsed_file}")

            # Validate and clean the result
            if isinstance(claude_result, dict):
                # Ensure all required keys exist
                required_keys = ["vendor_info", "invoice_details", "line_items",
                               "taxes_fees", "payment_info", "compliance_flags"]
                for key in required_keys:
                    if key not in claude_result:
                        claude_result[key] = []

                print(f"Successfully converted Landing AI results - found "
                      f"{len(claude_result['vendor_info'])} vendor info, "
                      f"{len(claude_result['invoice_details'])} invoice details, "
                      f"{len(claude_result['line_items'])} line items, "
                      f"{len(claude_result['taxes_fees'])} taxes/fees, "
                      f"{len(claude_result['payment_info'])} payment info, "
                      f"{len(claude_result['compliance_flags'])} compliance flags")
                return claude_result
            else:
                print(f"Claude result is not a dictionary: {type(claude_result)}")

        except json.JSONDecodeError as e:
            print(f"Error parsing Claude's JSON response: {e}")
            print(f"Raw response preview: {response.content[0].text[:500]}")

            # Try to fix truncated JSON by adding missing closing braces
            try:
                response_text = response.content[0].text.strip()

                # Strip markdown if present
                if response_text.startswith('```json'):
                    response_text = response_text[7:]
                if response_text.startswith('```'):
                    response_text = response_text[3:]
                if response_text.endswith('```'):
                    response_text = response_text[:-3]

                response_text = response_text.strip()

                # Count braces to detect truncation
                open_braces = response_text.count('{')
                close_braces = response_text.count('}')
                open_brackets = response_text.count('[')
                close_brackets = response_text.count(']')

                # Add missing closing characters
                missing_brackets = open_brackets - close_brackets
                missing_braces = open_braces - close_braces

                if missing_brackets > 0:
                    response_text += ']' * missing_brackets
                if missing_braces > 0:
                    response_text += '}' * missing_braces

                print(f"Attempting to fix truncated JSON by adding "
                      f"{missing_brackets} brackets and {missing_braces} braces")
                claude_result = json.loads(response_text)

                print("Successfully parsed fixed JSON!")
                return claude_result

            except Exception as fix_error:
                print(f"Failed to fix truncated JSON: {fix_error}")

            # Save the error for debugging
            error_file = os.path.join(logs_dir,
                                     f"{log_file_base}_json_parse_error.txt")
            with open(error_file, 'w') as f:
                f.write(f"JSON Parse Error: {str(e)}\n")
                f.write(f"Raw Response:\n{response.content[0].text}")
            print(f"Saved JSON parse error to: {error_file}")

        # Fallback: return empty structure
        print("Returning empty result structure due to conversion failure")
        return {
            "vendor_info": [],
            "invoice_details": [],
            "line_items": [],
            "taxes_fees": [],
            "payment_info": [],
            "compliance_flags": []
        }

    except Exception as e:
        print(f"Error converting Landing AI results: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")

        # Log the conversion error
        try:
            logs_dir = os.path.join(os.getcwd(), "logs")
            if not os.path.exists(logs_dir):
                os.makedirs(logs_dir)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            error_file = os.path.join(logs_dir,
                                     f"{timestamp}_conversion_error.txt")

            with open(error_file, 'w') as f:
                f.write(f"Error converting Landing AI results: {str(e)}\n")
                f.write(f"Traceback: {traceback.format_exc()}\n")
        except Exception as log_error:
            print(f"Failed to log conversion error: {log_error}")

        return {
            "vendor_info": [],
            "invoice_details": [],
            "line_items": [],
            "taxes_fees": [],
            "payment_info": [],
            "compliance_flags": []
        }

@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=4, max=10),
    reraise=True
)
def evaluate_and_correct_invoice_extraction(pdf_path, extraction_result,
                                          raw_markdown=None, raw_chunks=None):
    """
    Evaluates and corrects the invoice extraction result using Azure OpenAI
    Compares Claude's extraction against the raw Landing AI output
    """
    try:
        print(f"Evaluating invoice extraction from PDF: {pdf_path}")

        # Create logs directory if it doesn't exist
        logs_dir = os.path.join(os.getcwd(), "logs")
        if not os.path.exists(logs_dir):
            os.makedirs(logs_dir)

        # Generate a unique file name for this evaluation
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_name = os.path.basename(pdf_path)
        log_file_base = f"{timestamp}_{pdf_name}_eval"

        # Convert extraction_result to a string format for the prompt
        extraction_json = json.dumps(extraction_result, indent=2,
                                   cls=NumpyEncoder)

        # Create raw data context if available
        raw_data_context = ""
        if raw_markdown or raw_chunks:
            raw_data_context = f"""
RAW DOCUMENT DATA (Original Data):
MARKDOWN:
{raw_markdown or "No markdown data available"}

STRUCTURED CHUNKS:
{json.dumps(raw_chunks, indent=2, cls=NumpyEncoder) if raw_chunks
 else "No chunks data available"}

"""

        # Create the evaluation prompt for OpenAI
        prompt = f"""Evaluate and correct the extraction of invoice information
        from this invoice document. This may include vendor invoices,
        purchase orders, receipts, or billing statements.

TASK: Review the extracted data against the raw document output and ensure invoice
information is captured along with any compliance or processing flags.

{raw_data_context}
EXTRACTED DATA (To be evaluated):
{extraction_json}

Your job is to:
1. Compare the extraction against the raw document data above
2. Check if any invoice information that is clearly present in the raw data
   was missed
3. Verify that all invoice details were correctly extracted
4. Fix any errors in the extraction
5. Ensure ALL required fields are captured

REQUIRED INVOICE INFORMATION TO VERIFY:
1. VENDOR INFO: Name, Address, Contact, Tax ID, Payment Terms
2. INVOICE DETAILS: Invoice Number, Invoice Date, Due Date, PO Number, Currency
3. LINE ITEMS: Item descriptions, quantities, unit prices, line totals
4. TAXES & FEES: Tax amounts/rates, shipping, handling, discounts
5. PAYMENT INFO: Total due, amount paid, balance due, payment method
6. COMPLIANCE FLAGS: Duplicate detection, discrepancies, missing fields, tax issues

Carefully examine the following aspects:
1. VENDOR INFORMATION: Ensure vendor details are properly extracted
2. INVOICE HEADER: Verify invoice numbers, dates, and amounts are correct
3. LINE ITEMS: Check all line items are captured with correct quantities and prices
4. CALCULATIONS: Verify totals and tax calculations are accurate
5. COMPLIANCE: Check for any issues or flags that need attention
6. NUMBER FORMATS: Fix any issues with commas, decimal points, or currency
   symbols in numeric fields

IMPORTANT: Include all invoice information from vendor invoices, receipts, and billing
documents. Remove any unrelated transactions, instructions, explanations, or non-relevant data.

The response MUST maintain the exact same JSON structure as the input with
these exact keys:
- "vendor_info": Array of vendor information tables
- "invoice_details": Array of invoice detail tables
- "line_items": Array of line item tables
- "taxes_fees": Array of tax/fee tables
- "payment_info": Array of payment information tables
- "compliance_flags": Array of compliance flag tables

Each table should have a "data" property containing CSV data with
appropriate headers.

If any field cannot be determined, use "UNKNOWN" for that field, but ALL
fields MUST be present in the response.

Return ONLY valid JSON with NO explanatory text."""

        # Call Azure OpenAI API for evaluation
        try:
            # Initialize Azure OpenAI client with key-based authentication
            client = AzureOpenAI(
                azure_endpoint=AZURE_ENDPOINT,
                api_key=AZURE_API_KEY,
                api_version="2025-01-01-preview",
            )

            response = client.chat.completions.create(
                model=AZURE_DEPLOYMENT,
                messages=[
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                max_tokens=8192,
                temperature=0.1,
                response_format={"type": "json_object"}
            )
        except Exception as error:
            raise Exception(f"Error with Azure OpenAI model: {error}")

        # Save OpenAI response for debugging
        openai_response_file = os.path.join(logs_dir,
                                           f"{log_file_base}_openai_response.txt")
        with open(openai_response_file, 'w', encoding='utf-8') as f:
            f.write(response.choices[0].message.content
                   if response.choices[0].message.content else "EMPTY RESPONSE")
        print(f"Saved OpenAI response to: {openai_response_file}")

        # Parse and validate the response
        response_text = response.choices[0].message.content.strip()

        if not response_text:
            print("Warning: OpenAI returned empty response, using original extraction")
            return extraction_result

        # Strip markdown code blocks if present
        if response_text.startswith('```json'):
            response_text = response_text[7:]  # Remove ```json
        if response_text.startswith('```'):
            response_text = response_text[3:]   # Remove ```
        if response_text.endswith('```'):
            response_text = response_text[:-3]  # Remove closing ```

        response_text = response_text.strip()

        try:
            corrected_result = json.loads(response_text)
        except json.JSONDecodeError as parse_error:
            print(f"Error parsing OpenAI JSON response: {parse_error}")
            print(f"Response text preview: {response_text[:500]}")

            # Try to fix truncated JSON
            try:
                # Count braces to detect truncation
                open_braces = response_text.count('{')
                close_braces = response_text.count('}')
                open_brackets = response_text.count('[')
                close_brackets = response_text.count(']')

                # Add missing closing characters
                missing_brackets = open_brackets - close_brackets
                missing_braces = open_braces - close_braces

                if missing_brackets > 0:
                    response_text += ']' * missing_brackets
                if missing_braces > 0:
                    response_text += '}' * missing_braces

                print(f"Attempting to fix truncated OpenAI JSON by adding "
                      f"{missing_brackets} brackets and {missing_braces} braces")
                corrected_result = json.loads(response_text)
                print("Successfully parsed fixed OpenAI JSON!")

            except Exception as fix_error:
                print(f"Failed to fix truncated OpenAI JSON: {fix_error}")
                print("Using original extraction result due to parsing failure")
                return extraction_result

        # Log the evaluation result
        eval_result_file = os.path.join(logs_dir,
                                       f"{log_file_base}_result.json")
        with open(eval_result_file, 'w', encoding='utf-8') as f:
            json.dump(corrected_result, f, indent=2, cls=NumpyEncoder, ensure_ascii=False)

        # Verify the corrected result has the required structure
        if not isinstance(corrected_result, dict):
            print("Warning: Evaluation didn't return a dictionary, using original extraction")
            return extraction_result

        # Ensure the corrected result has all required keys
        required_keys = ["vendor_info", "invoice_details", "line_items",
                        "taxes_fees", "payment_info", "compliance_flags"]
        for key in required_keys:
            if key not in corrected_result:
                print(f"Warning: Evaluation result missing '{key}', "
                      f"adding from original extraction")
                corrected_result[key] = extraction_result.get(key, [])

        return corrected_result

    except Exception as e:
        print(f"Error evaluating extraction: {e}")
        # Log the error
        try:
            logs_dir = os.path.join(os.getcwd(), "logs")
            if not os.path.exists(logs_dir):
                os.makedirs(logs_dir)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            error_file = os.path.join(logs_dir,
                                     f"{timestamp}_eval_error.txt")

            with open(error_file, 'w') as f:
                f.write(f"Error evaluating extraction: {str(e)}")
        except Exception as log_error:
            print(f"Failed to log error: {log_error}")

        # Return the original extraction result if evaluation fails
        return extraction_result


def create_invoice_analysis_excel(invoice_data, output_path):
    """
    Create a comprehensive Excel report from the invoice analysis data
    """
    try:
        # Create a new workbook
        wb = Workbook()

        # Remove the default sheet
        wb.remove(wb.active)

        # Define sections with their sheet names
        sections = [
            ('vendor_info', 'Vendor Information'),
            ('invoice_details', 'Invoice Details'),
            ('line_items', 'Line Items'),
            ('taxes_fees', 'Taxes & Fees'),
            ('payment_info', 'Payment Information'),
            ('compliance_flags', 'Compliance Flags and Risk Factors')
        ]

        # Create sheets for each section
        for section_key, sheet_name in sections:
            if invoice_data.get(section_key):
                ws = wb.create_sheet(title=sheet_name)

                # Add header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")

                row_num = 1

                for table_data in invoice_data[section_key]:
                    if isinstance(table_data, dict) and 'data' in table_data:
                        try:
                            # Parse CSV data manually
                            lines = table_data['data'].strip().split('\n')
                            if len(lines) > 1:
                                # Add headers
                                headers = lines[0].split(',')
                                for col_num, header in enumerate(headers, 1):
                                    cell = ws.cell(row=row_num, column=col_num, value=header.strip())
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = header_alignment

                                row_num += 1

                                # Add data rows
                                for line in lines[1:]:
                                    if ',' in line:
                                        # Split only on first comma to handle commas in values
                                        first_comma = line.find(',')
                                        if first_comma != -1:
                                            field_name = line[:first_comma].strip()
                                            field_value = line[first_comma+1:].strip()

                                            if field_value and field_value != 'UNKNOWN':
                                                ws.cell(row=row_num, column=1, value=field_name)
                                                ws.cell(row=row_num, column=2, value=field_value)
                                                row_num += 1

                                # Add spacing between tables
                                row_num += 2

                        except Exception as e:
                            print(f"Error processing {section_key}: {e}")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

        # Create summary sheet
        summary_ws = wb.create_sheet(title="Summary", index=0)
        summary_ws['A1'] = "Invoice Analysis Report"
        summary_ws['A1'].font = Font(bold=True, size=16)
        summary_ws['A3'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        summary_ws['A5'] = "Executive Summary"
        summary_ws['A5'].font = Font(bold=True, size=14)
        summary_ws['A6'] = ("This report provides a comprehensive analysis of the invoice "
                           "documents, including vendor information, invoice details, line items, "
                           "and key terms and conditions.")

        # Auto-adjust summary column width
        summary_ws.column_dimensions['A'].width = 80

        # Save the workbook
        wb.save(output_path)

        # Verify the file was created successfully
        if os.path.exists(output_path) and os.path.getsize(output_path) > 1000:
            print(f"✅ Invoice analysis Excel saved successfully: {output_path} ({os.path.getsize(output_path)} bytes)")
        else:
            raise Exception(f"Excel file was not created properly or is too small")

    except Exception as e:
        print(f"❌ Error creating Excel report: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")

        # Fallback: create a simple text file
        try:
            text_path = output_path.replace('.xlsx', '.txt')
            with open(text_path, 'w', encoding='utf-8') as f:
                f.write("Invoice Analysis Report\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Error: {str(e)}\n")
                f.write("Excel generation failed, please check the logs.\n\n")

                # Add some basic invoice data if available
                if invoice_data:
                    f.write("Available Data:\n")
                    for key, value in invoice_data.items():
                        f.write(f"{key}: {value}\n")

            print(f"📝 Created fallback text file: {text_path}")
        except Exception as fallback_error:
            print(f"❌ Fallback text file creation also failed: {fallback_error}")

            # Process each section and write to text file
            sections = [
                ('vendor_info', '1. Vendor Information'),
                ('invoice_details', '2. Invoice Details'),
                ('line_items', '3. Line Items'),
                ('taxes_fees', '4. Taxes & Fees'),
                ('payment_info', '5. Payment Information'),
                ('compliance_flags', '6. Compliance Flags and Risk Factors')
            ]

            for section_key, section_title in sections:
                if invoice_data.get(section_key):
                    with open(text_path, 'a') as f:
                        f.write(f"{section_title}\n")
                        f.write("-" * len(section_title) + "\n")

                        for table_data in invoice_data[section_key]:
                            if isinstance(table_data, dict) and 'data' in table_data:
                                try:
                                    lines = table_data['data'].strip().split('\n')
                                    if len(lines) > 1:
                                        for line in lines[1:]:
                                            if ',' in line:
                                                first_comma = line.find(',')
                                                field_name = line[:first_comma].strip()
                                                field_value = line[first_comma+1:].strip()
                                                if field_value and field_value != 'UNKNOWN':
                                                    f.write(f"{field_name}: {field_value}\n")
                                        f.write("\n")
                                except Exception as e:
                                    print(f"Error processing {section_key}: {e}")

            print(f"Text report saved to: {text_path}")
        except Exception as fallback_error:
            print(f"Fallback text file creation also failed: {fallback_error}")


def process_invoice_document_with_landing_ai(pdf_path, output_dir):
    """
    Process invoice document using Landing AI as the first layer, then apply
    validation layers.
    """
    print(f"Processing invoice document with Landing AI: {pdf_path}")

    try:
        # Step 1: Use Landing AI to extract data from entire PDF
        result = extract_invoice_data_with_landing_ai(pdf_path)

        if not result:
            print("No data extracted from Landing AI")
            return {
                "vendor_info": [],
                "invoice_details": [],
                "line_items": [],
                "taxes_fees": [],
                "payment_info": [],
                "compliance_flags": []
            }

        # Step 2: Apply evaluation/correction layer (Azure OpenAI)
        print("Running evaluator to validate and correct extraction...")
        # Extract raw data to pass to evaluator
        raw_markdown = None
        raw_chunks = None
        if hasattr(result, '_raw_markdown'):
            raw_markdown = result._raw_markdown
        if hasattr(result, '_raw_chunks'):
            raw_chunks = result._raw_chunks

        corrected_result = evaluate_and_correct_invoice_extraction(
            pdf_path, result, raw_markdown, raw_chunks)

        return corrected_result

    except Exception as e:
        print(f"Error processing invoice document with Landing AI: {e}")

        # Check if it's an encryption error
        if "encrypted" in str(e).lower() or "password" in str(e).lower():
            print("PDF is encrypted/password-protected - cannot process this file")
            # Return a clear error response that will be sent to the frontend
            raise Exception("PDF_ENCRYPTED: PDF file is encrypted or password-protected. Cannot process encrypted PDFs. Please provide an unprotected PDF file.")

        return {
            "vendor_info": [],
            "invoice_details": [],
            "line_items": [],
            "taxes_fees": [],
            "payment_info": [],
            "compliance_flags": []
        }


# Include routers
app.include_router(auth.router, prefix="/auth", tags=["Authentication"])
app.include_router(invoices.router, prefix="/api/v1", tags=["Invoices"])


@app.post("/parse-invoices")
async def parse_invoice_documents(
    pdf_files: list[UploadFile] = File(...),
    invoice_id: str = Form(None)
):
    """
    Main endpoint that uses Landing AI to parse invoice documents and extract
    comprehensive invoice information using the AI pipeline.
    """
    global last_processed_file

    if not pdf_files or len(pdf_files) < 1:
        return {"error": "At least 1 PDF file is required"}

    print("Processing invoice documents with Landing AI AI pipeline...")

    # Add memory limit check
    total_size = 0
    MAX_TOTAL_SIZE = 100 * 1024 * 1024  # 100MB limit

    for pdf_file in pdf_files:
        # Read file in chunks to check size
        size = 0
        while chunk := await pdf_file.read(8192):
            size += len(chunk)
            if size > MAX_TOTAL_SIZE:
                return {"error": f"File {pdf_file.filename} is too large"}
        total_size += size
        await pdf_file.seek(0)  # Reset file pointer

        if not pdf_file.filename.endswith('.pdf'):
            return {"error": f"File {pdf_file.filename} is not a PDF file"}

    if total_size > MAX_TOTAL_SIZE:
        return {"error": "Total file size too large"}

    print("Total file size is within the limit")

    # Create temporary directories
    temp_dir = "./tmp"
    output_dir = os.path.join(temp_dir, str(uuid.uuid4()))
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, "invoice_analysis.xlsx")

    try:
        # Initialize combined invoice data
        combined_invoice_data = {
            "vendor_info": [],
            "invoice_details": [],
            "line_items": [],
            "taxes_fees": [],
            "payment_info": [],
            "compliance_flags": []
        }

        # Process all uploaded documents
        documents = []
        for i, pdf_file in enumerate(pdf_files):
            doc_name = f"Document {i+1}"
            if pdf_file.filename:
                doc_name = f"Document {i+1}: {pdf_file.filename}"
            documents.append((doc_name, pdf_file))

        # Process all documents and collect comprehensive data
        all_document_data = []

        for doc_name, pdf_file in documents:
            try:
                print(f"Processing {doc_name} with Landing AI AI pipeline")
                gc.collect()

                # Use provided invoice_id or generate a new one if not provided
                if not invoice_id:
                    invoice_id = str(uuid.uuid4())
                    print(f"⚠️ No invoice_id provided, generated new one: {invoice_id}")
                else:
                    print(f"✅ Using provided invoice_id: {invoice_id}")

                # Save original file to originals directory
                originals_dir = os.path.join(os.getcwd(), "originals")
                if not os.path.exists(originals_dir):
                    os.makedirs(originals_dir)

                original_file_path = os.path.join(originals_dir, f"{invoice_id}_{pdf_file.filename}")
                with open(original_file_path, "wb") as f:
                    content = await pdf_file.read()
                    f.write(content)

                # Create a copy for processing
                pdf_path = os.path.join(temp_dir, f"input_{pdf_file.filename}")
                with open(pdf_path, "wb") as f:
                    f.write(content)

                # Process PDF using Landing AI + AI validation pipeline
                result = process_invoice_document_with_landing_ai(pdf_path, output_dir)

                if result:
                    # Store document-specific data for comprehensive analysis
                    all_document_data.append({
                        'document_name': doc_name,
                        'data': result
                    })

                    # Also combine data for immediate use - preserve all data from both documents
                    categories = ["vendor_info", "invoice_details", "line_items",
                                 "taxes_fees", "payment_info", "compliance_flags"]

                    for category in categories:
                        if result.get(category):
                            # Add data from this document to the combined data
                            # This ensures we preserve data from both documents
                            combined_invoice_data[category].extend(result[category])
                            print(f"Added {len(result[category])} items to {category} from {doc_name}")

                # Clean up PDF file
                os.remove(pdf_path)

            except Exception as e:
                print(f"Error processing PDF {pdf_file.filename}: {e}")

                # Check if it's an encryption error and provide specific feedback
                if "encrypted" in str(e).lower() or "password" in str(e).lower():
                    print(f"⚠️  PDF {pdf_file.filename} is password-protected - skipping this file")
                    # Add error information to the combined data
                    combined_invoice_data["errors"] = combined_invoice_data.get("errors", [])
                    combined_invoice_data["errors"].append({
                        "file": pdf_file.filename,
                        "error": "PDF is encrypted or password-protected",
                        "details": "Cannot process encrypted PDFs. Please provide an unprotected PDF file."
                    })
                else:
                    # Add other errors as well
                    combined_invoice_data["errors"] = combined_invoice_data.get("errors", [])
                    combined_invoice_data["errors"].append({
                        "file": pdf_file.filename,
                        "error": str(e),
                        "details": "Processing failed for this file"
                    })

                continue

        # If we have multiple documents, perform cross-document validation
        if len(all_document_data) > 1:
            print(f"Performing cross-document validation for {len(all_document_data)} documents")

            # Create a comprehensive analysis prompt for multiple documents
            cross_doc_prompt = f"""You have processed {len(all_document_data)} invoice documents - combine and validate the extracted data.

DOCUMENTS PROCESSED:
"""

            for i, doc_data in enumerate(all_document_data):
                cross_doc_prompt += f"\nDocument {i+1}: {doc_data['document_name']}\n"
                cross_doc_prompt += f"Data: {json.dumps(doc_data['data'], indent=2, cls=NumpyEncoder)}\n"

            cross_doc_prompt += f"""

CURRENT COMBINED DATA:
{json.dumps(combined_invoice_data, indent=2, cls=NumpyEncoder)}

TASK:
1. Combine data from multiple invoice documents
2. Remove duplicates and conflicting information
3. Validate totals and calculations across documents
4. Flag any discrepancies between documents
5. Create a consolidated view of all invoice information

VALIDATION RULES:
- If the same vendor appears in multiple documents, combine their information
- If invoice numbers conflict, flag as potential duplicate
- Sum totals from multiple documents if they represent different invoices
- Flag any calculation discrepancies
- Preserve all unique line items and vendor information

Return a consolidated JSON with the same structure but with validated, combined data.
"""

            try:
                # Use Claude to validate cross-document data
                client = anthropic.Anthropic()
                response = client.messages.create(
                    model="claude-3-7-sonnet-latest",
                    max_tokens=8192,
                    messages=[{
                        "role": "user",
                        "content": cross_doc_prompt
                    }]
                )

                # Parse and use the cross-document validation result
                response_text = response.content[0].text.strip()

                # Strip markdown if present
                if response_text.startswith('```json'):
                    response_text = response_text[7:]
                if response_text.startswith('```'):
                    response_text = response_text[3:]
                if response_text.endswith('```'):
                    response_text = response_text[:-3]

                response_text = response_text.strip()

                validated_result = json.loads(response_text)

                # Update combined data with validated result
                if isinstance(validated_result, dict):
                    for key in combined_invoice_data.keys():
                        if key in validated_result:
                            combined_invoice_data[key] = validated_result[key]

                    print("Cross-document validation completed successfully")

            except Exception as cross_error:
                print(f"Cross-document validation failed: {cross_error}")
                # Continue with original combined data

        # Create Excel report from combined data
        if any(combined_invoice_data.values()):
            # Create exports directory
            exports_dir = os.path.join(os.getcwd(), "exports")
            if not os.path.exists(exports_dir):
                os.makedirs(exports_dir)

            excel_path = os.path.join(exports_dir, "invoice_analysis.xlsx")
            create_invoice_analysis_excel(combined_invoice_data, excel_path)

            # Check if Excel was created successfully
            if os.path.exists(excel_path) and os.path.getsize(excel_path) > 0:
                last_processed_file = excel_path

                # Return file response with invoice ID in headers for correlation
                response = FileResponse(
                    excel_path,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename="invoice_analysis.xlsx"
                )
                # Add invoice ID to response headers for Next.js to correlate
                response.headers["X-Invoice-ID"] = invoice_id
                
                # Update database directly with extracted data
                update_invoice_database_directly(invoice_id, combined_invoice_data, pdf_files[0].filename)
                
                return response
            # Check if text file was created as fallback
            elif os.path.exists(excel_path.replace('.xlsx', '.txt')):
                text_path = excel_path.replace('.xlsx', '.txt')
                last_processed_file = text_path

                return FileResponse(
                    text_path,
                    media_type="text/plain",
                    filename="invoice_analysis.txt"
                )

        return {"error": "No valid invoice data was found in the PDFs"}

    except Exception as e:
        return {"error": f"Error processing files: {str(e)}"}


@app.get("/last-processed-invoices")
async def get_last_processed_invoice():
    """Returns the last successfully processed Excel file from invoice parsing endpoint"""
    global last_processed_file

    if last_processed_file and os.path.exists(last_processed_file):
        # Determine the correct media type and filename based on file extension
        if last_processed_file.endswith('.xlsx'):
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = "invoice_analysis.xlsx"
        elif last_processed_file.endswith('.txt'):
            media_type = "text/plain"
            filename = "invoice_analysis.txt"
        else:
            media_type = "application/octet-stream"
            filename = "invoice_analysis"

        return FileResponse(
            last_processed_file,
            media_type=media_type,
            filename=filename
        )
    else:
        return {"error": "No processed file available. Please process invoice documents first using /parse-invoices endpoint."}


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "service": "Invoice Processing Service",
        "version": "1.0.0"
    }

@app.post("/api/v1/invoices/upload", response_model=InvoiceResponse)
async def upload_invoice(
    file: UploadFile = File(...),
    current_user: str = Depends(verify_token),
    db: Session = Depends(get_db)
):
    """
    Upload and process an invoice
    - Parse invoice using Landing AI
    - Classify fields using Claude/OpenAI
    - Store in database with embeddings
    """
    try:
        # Process the invoice
        result = await invoice_processor.process_invoice(file, current_user)

        # Store in database
        invoice = Invoice(
            user_id=current_user,
            filename=file.filename,
            extracted_data=result["extracted_data"],
            classified_data=result["classified_data"],
            embeddings=result["embeddings"],
            status="processed"
        )

        db.add(invoice)
        db.commit()
        db.refresh(invoice)

        return InvoiceResponse.from_orm(invoice)

    except Exception as e:
        logger.error("Invoice processing failed", error=str(e), user=current_user)
        raise HTTPException(status_code=500, detail="Invoice processing failed")

@app.post("/api/v1/invoices/search", response_model=List[InvoiceSearchResponse])
async def search_invoices(
    query: InvoiceQuery,
    current_user: str = Depends(verify_token),
    db: Session = Depends(get_db)
):
    """
    Semantic search through processed invoices using RAG
    """
    try:
        results = await search_service.search_invoices(
            query.query,
            current_user,
            limit=query.limit,
            threshold=query.threshold
        )

        return [
            InvoiceSearchResponse(
                invoice_id=result["invoice_id"],
                filename=result["filename"],
                relevance_score=result["score"],
                matched_content=result["content"],
                extracted_data=result["extracted_data"]
            )
            for result in results
        ]

    except Exception as e:
        logger.error("Invoice search failed", error=str(e), user=current_user)
        raise HTTPException(status_code=500, detail="Search failed")

@app.get("/api/v1/invoices/{invoice_id}/export")
async def export_invoice(
    invoice_id: str,
    current_user: str = Depends(verify_token),
    db: Session = Depends(get_db)
):
    """
    Export invoice data to Excel and upload to S3
    """
    try:
        # Get invoice from database
        invoice = db.query(Invoice).filter(
            Invoice.id == invoice_id,
            Invoice.user_id == current_user
        ).first()

        if not invoice:
            raise HTTPException(status_code=404, detail="Invoice not found")

        # Export to Excel and upload to S3
        export_url = await invoice_processor.export_to_excel(invoice)

        return {"export_url": export_url, "status": "success"}

    except Exception as e:
        logger.error("Invoice export failed", error=str(e), user=current_user)
        raise HTTPException(status_code=500, detail="Export failed")

@app.get("/last-processed-invoices")
async def get_last_processed_invoices():
    """
    Get the last processed invoice Excel file
    """
    global last_processed_file

    if not last_processed_file or not os.path.exists(last_processed_file):
        return {"error": "No processed invoices found"}

    # Return the file
    if last_processed_file.endswith('.xlsx'):
        return FileResponse(
            last_processed_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="invoice_analysis.xlsx"
        )
    else:
        return FileResponse(
            last_processed_file,
            media_type="text/plain",
            filename="invoice_analysis.txt"
        )

@app.get("/invoices/{invoice_id}/original")
async def get_original_invoice(invoice_id: str):
    """
    Get the original uploaded invoice file
    """
    # Create originals directory if it doesn't exist
    originals_dir = os.path.join(os.getcwd(), "originals")
    if not os.path.exists(originals_dir):
        os.makedirs(originals_dir)

    # Look for the original file
    original_file_path = None
    for filename in os.listdir(originals_dir):
        if filename.startswith(invoice_id) and filename.endswith('.pdf'):
            original_file_path = os.path.join(originals_dir, filename)
            break

    if not original_file_path or not os.path.exists(original_file_path):
        return {"error": "Original file not found"}

    # Return the file
    return FileResponse(
        original_file_path,
        media_type="application/pdf",
        filename=os.path.basename(original_file_path)
    )

@app.on_event("startup")
async def startup_event():
    """Initialize services on startup"""
    logger.info("Starting Invoice Processing Service")

@app.on_event("shutdown")
async def shutdown_event():
    """Cleanup on shutdown"""
    logger.info("Shutting down Invoice Processing Service")

if __name__ == "__main__":
    uvicorn.run(
        "app.main:app",
        host=os.getenv("HOST", "0.0.0.0"),
        port=settings.PORT,
        reload=settings.DEBUG,
        log_level="info"
    )
